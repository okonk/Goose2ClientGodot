"""Reads the three workbooks + mapping tables, writes the remapped workbook.

Env overrides (defaults shown):
  ASP_XLSX=/home/hayden/Downloads/Aspereta Goose Data.xlsx
  XEN_XLSX=/home/hayden/Downloads/Xendria Aspereta Goose Data.xlsx
  ILL_XLSX=/home/hayden/Downloads/Illutia Goose Data.xlsx
  OUT_XLSX=/home/hayden/Downloads/Aspereta Goose Data (Illutia).xlsx
"""
import os
import openpyxl
from remap.sheets import Sheet
from remap.core import Remapper
from remap.mappings import load_item_mapping, load_graphics_mapping
from remap.illutia_index import build_display_tile_index
from remap.items import transform_items
from remap.npcs import transform_npcs
from remap.spells import transform_spells, transform_spell_effects
from remap.merge import rename_map_files, merge_xendria_npcs, XENDRIA_COPY_SHEETS

REPO = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
DEFAULTS = {
    "ASP_XLSX": "/home/hayden/Downloads/Aspereta Goose Data.xlsx",
    "XEN_XLSX": "/home/hayden/Downloads/Xendria Aspereta Goose Data.xlsx",
    "ILL_XLSX": "/home/hayden/Downloads/Illutia Goose Data.xlsx",
    "OUT_XLSX": "/home/hayden/Downloads/Aspereta Goose Data (Illutia).xlsx",
    "ITEM_MAP": os.path.join(REPO, "tools/AssetConverter/data/aspereta-item-mapping.tsv"),
    "GFX_MAP": os.path.join(REPO, "tools/AssetConverter/data/aspereta-mapping.tsv"),
}


def cfg(key):
    return os.environ.get(key, DEFAULTS[key])


def read_book(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    sheets = {}
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(values_only=True))
        header = list(rows[0]) if rows else []
        data = [list(r) for r in rows[1:] if r and any(v is not None for v in r)]
        sheets[name] = Sheet(name, header, data)
    wb.close()
    return sheets


def write_book(path, ordered_sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in ordered_sheets:
        ws = wb.create_sheet(sheet.name)
        ws.append(sheet.header)
        for row in sheet.rows:
            ws.append(row)
    wb.save(path)


def run():
    asp = read_book(cfg("ASP_XLSX"))
    xen = read_book(cfg("XEN_XLSX"))
    ill = read_book(cfg("ILL_XLSX"))

    remapper = Remapper(load_graphics_mapping(cfg("GFX_MAP")),
                        load_item_mapping(cfg("ITEM_MAP")))
    ill_index = build_display_tile_index(ill["Items"])

    # 1. Xendria merge (before transforms so merged NPCs get remapped too)
    merge_xendria_npcs(asp["NPCs"], xen["NPCs"], asp["NPC Spawns"], xen["NPC Spawns"],
                       asp["Maps"], xen["Maps"], remapper.warnings)
    for name in XENDRIA_COPY_SHEETS:
        asp[name] = xen[name]

    # 2. Transforms
    transform_items(asp["Items"], remapper, ill_index)
    transform_npcs(asp["NPCs"], remapper)
    transform_spells(asp["Spells"], remapper)
    transform_spell_effects(asp["Spell Effects"], remapper)
    rename_map_files(asp["Maps"])

    # 3. Referential validation (warnings only)
    item_ids = {int(float(r[0])) for r in asp["Items"].rows}
    npc_ids = {int(float(r[0])) for r in asp["NPCs"].rows}
    spell_ids = {int(float(r[0])) for r in asp["Spells"].rows}
    for r in asp["Quest Rewards"].rows:
        v = asp["Quest Rewards"].get(r, "long value")
        if str(asp["Quest Rewards"].get(r, "reward type")).lower().startswith("item") \
                and v and int(float(v)) not in item_ids:
            remapper.warn(f"Quest Rewards quest={r[1]}: item {v} not in Items")
    for r in asp["Class Levelup Spells"].rows:
        v = asp["Class Levelup Spells"].get(r, "spell id")
        if v and int(float(v)) not in spell_ids:
            remapper.warn(f"Class Levelup Spells: spell {v} not in Spells")
    for r in asp["NPC Spawns"].rows:
        v = asp["NPC Spawns"].get(r, "npc id")
        if v and int(float(v)) not in npc_ids:
            remapper.warn(f"NPC Spawns: npc {v} not in NPCs")

    # 4. Write output + warnings, preserving the base workbook's sheet order
    order = list(asp.values())
    out = cfg("OUT_XLSX")
    write_book(out, order)
    warn_path = os.path.splitext(out)[0] + ".warnings.txt"
    with open(warn_path, "w") as f:
        f.write("\n".join(remapper.warnings))
    print(f"wrote {out}")
    print(f"{len(remapper.warnings)} warnings -> {warn_path}")


if __name__ == "__main__":
    run()
