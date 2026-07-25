"""End-to-end run against the real workbooks. Skips cleanly if inputs are absent
(e.g. CI); on Hayden's machine it must run."""
import os, subprocess, sys, tempfile
import openpyxl
import pytest
from remap.main import DEFAULTS

pytestmark = pytest.mark.skipif(
    not (os.path.exists(DEFAULTS["ASP_XLSX"]) and os.path.exists(DEFAULTS["GFX_MAP"])),
    reason="real inputs not present")

def test_full_run_invariants(tmp_path):
    out = str(tmp_path / "out.xlsx")
    env = dict(os.environ, OUT_XLSX=out)
    subprocess.run([sys.executable, "-m", "remap.main"],
                   cwd=os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   env=env, check=True)
    wb = openpyxl.load_workbook(out, read_only=True)

    maps = list(wb["Maps"].iter_rows(values_only=True))
    fn_col = [i for i, h in enumerate(maps[0]) if h == "filename"][0]
    for r in maps[1:]:
        if r[fn_col]:
            assert str(r[fn_col]).startswith("Map1"), r[fn_col]   # Map10001.map etc.

    npcs = list(wb["NPCs"].iter_rows(values_only=True))
    # 182 Aspereta + 8 quest givers (311–318 renumbered to 183–190)
    assert len(npcs) - 1 == 190
    npc_ids = sorted(int(float(r[0])) for r in npcs[1:] if r and r[0] is not None)
    assert npc_ids[-8:] == list(range(183, 191))

    quests = list(wb["Quests"].iter_rows(values_only=True))
    assert len(quests) - 1 > 0           # xendria quests copied

    items = list(wb["Items"].iter_rows(values_only=True))
    hdr = [str(h) if h else "" for h in items[0]]
    gt = hdr.index("graphic tile"); gf = next(i for i, h in enumerate(hdr) if h.startswith("graphic file"))
    for r in items[1:]:
        if r[gt]:
            # every remapped tile has a sheet: illutia donor or injected 20000+
            assert r[gf] is not None and int(r[gf]) > 0, (r[0], r[gt], r[gf])
    wb.close()
